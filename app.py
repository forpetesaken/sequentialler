import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import Rule, ColorScaleRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
import csv
import io
import re

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Alignment Formatter",
    page_icon="🧬",
    layout="wide",
)

# ── Styling ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { background: #0f1117; }
    h1 { color: #e8eaf6; }
    h2, h3 { color: #c5cae9; }
    .stButton > button {
        background: #3949ab;
        color: white;
        border: none;
        border-radius: 6px;
        padding: 0.5rem 1.5rem;
        font-weight: 600;
    }
    .stButton > button:hover { background: #5c6bc0; }
    .domain-row {
        display: flex; align-items: center; gap: 8px;
        background: #1e2130; border-radius: 8px;
        padding: 8px 12px; margin-bottom: 6px;
    }
    .info-box {
        background: #1a237e22; border-left: 4px solid #3949ab;
        padding: 12px 16px; border-radius: 4px; margin: 12px 0;
    }
    div[data-testid="stColorPicker"] label { font-size: 0.75rem; }
</style>
""", unsafe_allow_html=True)

# ── Default amino acid colors ─────────────────────────────────────────────────
DEFAULT_AA_COLORS = {
    'A': '#E0E0E0', 'C': '#E0E0E0',
    'D': '#FFC7CE', 'E': '#FFC7CE',
    'F': '#FFEB9C', 'G': '#F9E6FE',
    'H': '#F8F6DA', 'I': '#FCC8C8',
    'K': '#D9E1F2', 'L': '#F7E6D9',
    'M': '#F3FBD5', 'N': '#CEF6E1',
    'P': '#D5DAFB', 'Q': '#DBF5EC',
    'R': '#FCC8E1', 'S': '#BABCFE',
    'T': '#FFFFC9', 'V': '#FCECC4',
    'W': '#99CCFF', 'Y': '#FCE4D6',
    '-': '#FFFFFF',
}

AA_GROUPS = {
    "Nonpolar / Small":   ['A', 'G', 'V', 'L', 'I', 'P'],
    "Polar uncharged":    ['S', 'T', 'C', 'M', 'N', 'Q'],
    "Aromatic":           ['F', 'W', 'Y', 'H'],
    "Positively charged": ['K', 'R'],
    "Negatively charged": ['D', 'E'],
    "Gap":                ['-'],
}

COLOR_TEMPLATES = {
    "Biochemical Groups": {
        'A': '#D8E2DC', 'G': '#D8E2DC', 'V': '#D8E2DC', 'L': '#D8E2DC', 'I': '#D8E2DC', 'P': '#D8E2DC',
        'S': '#D9F0FF', 'T': '#D9F0FF', 'C': '#D9F0FF', 'M': '#D9F0FF', 'N': '#D9F0FF', 'Q': '#D9F0FF',
        'F': '#FFE8CC', 'W': '#FFE8CC', 'Y': '#FFE8CC', 'H': '#FFE8CC',
        'K': '#F8D7FF', 'R': '#F8D7FF',
        'D': '#FFD6D6', 'E': '#FFD6D6',
        '-': '#FFFFFF',
    },
    "Ocean Breeze": {
        'A': '#E3F2FD', 'C': '#E0F7FA',
        'D': '#FFCDD2', 'E': '#F8BBD0',
        'F': '#FFE0B2', 'G': '#E1F5FE',
        'H': '#E8EAF6', 'I': '#D0F4DE',
        'K': '#D1C4E9', 'L': '#C8E6C9',
        'M': '#B2DFDB', 'N': '#DCEDC8',
        'P': '#BBDEFB', 'Q': '#C5E1A5',
        'R': '#E1BEE7', 'S': '#B3E5FC',
        'T': '#B2EBF2', 'V': '#C8E6C9',
        'W': '#FFF9C4', 'Y': '#FFE082',
        '-': '#FFFFFF',
    },
    "Sunset Sorbet": {
        'A': '#FFF3E0', 'C': '#FFE0B2',
        'D': '#FFCDD2', 'E': '#EF9A9A',
        'F': '#FFE082', 'G': '#F8BBD0',
        'H': '#F3E5F5', 'I': '#FFECB3',
        'K': '#E1BEE7', 'L': '#FFD180',
        'M': '#FFCCBC', 'N': '#FCE4EC',
        'P': '#FFAB91', 'Q': '#F8BBD0',
        'R': '#CE93D8', 'S': '#FFCC80',
        'T': '#FFE0B2', 'V': '#FFB74D',
        'W': '#FFF59D', 'Y': '#FFE082',
        '-': '#FFFFFF',
    },
    "Forest Mist": {
        'A': '#E8F5E9', 'C': '#D0F0C0',
        'D': '#FFEBEE', 'E': '#FDE2E4',
        'F': '#F1F8E9', 'G': '#E0F2F1',
        'H': '#E8F5E9', 'I': '#C8E6C9',
        'K': '#D7CCC8', 'L': '#A5D6A7',
        'M': '#B2DFDB', 'N': '#DCEDC8',
        'P': '#C5E1A5', 'Q': '#E6EE9C',
        'R': '#BCAAA4', 'S': '#C8E6C9',
        'T': '#DCE775', 'V': '#AED581',
        'W': '#F0F4C3', 'Y': '#FFF9C4',
        '-': '#FFFFFF',
    },
}

# ── Session state init ────────────────────────────────────────────────────────
if 'aa_colors' not in st.session_state:
    st.session_state.aa_colors = DEFAULT_AA_COLORS.copy()
if 'domains' not in st.session_state:
    st.session_state.domains = [
        {"start": 226, "end": 228,  "label": "YXF"},
        {"start": 244, "end": 244,  "label": "244"},
        {"start": 250, "end": 250,  "label": "250"},
        {"start": 253, "end": 253,  "label": "253"},
        {"start": 256, "end": 256,  "label": "256"},
        {"start": 258, "end": 260,  "label": "258-260"},
        {"start": 266, "end": 288,  "label": "ZF1"},
        {"start": 294, "end": 316,  "label": "ZF2"},
        {"start": 322, "end": 345,  "label": "ZF3"},
        {"start": 351, "end": 373,  "label": "ZF4"},
        {"start": 379, "end": 401,  "label": "ZF5"},
        {"start": 407, "end": 430,  "label": "ZF6"},
        {"start": 437, "end": 460,  "label": "ZF7"},
        {"start": 467, "end": 489,  "label": "ZF8"},
        {"start": 495, "end": 517,  "label": "ZF9"},
        {"start": 523, "end": 546,  "label": "ZF10"},
    ]

for aa, color in st.session_state.aa_colors.items():
    picker_key = f"color_{aa}"
    if picker_key not in st.session_state:
        st.session_state[picker_key] = color

# ── Helper: hex → openpyxl ARGB ──────────────────────────────────────────────
def hex_to_argb(hex_color: str) -> str:
    h = hex_color.lstrip('#')
    if len(h) == 6:
        return 'FF' + h.upper()
    return h.upper()

def apply_color_scheme(color_map):
    st.session_state.aa_colors = color_map.copy()
    for aa, color in color_map.items():
        st.session_state[f"color_{aa}"] = color

def render_alignment_preview(aa_colors):
    preview_rows = [
        ("Reference", "ACDEFGHIKLMNPQRSTVWY"),
        ("Variant A", "ACD-FGHIKLMNPQKSTVWY"),
        ("Variant B", "ACNEYGHI-LMNPQRSTAWY"),
        ("Variant C", "ACDEWGHIKLM-PQRSAV-Y"),
    ]

    row_html = []
    for name, seq in preview_rows:
        cells = "".join(
            f'<span style="display:inline-flex;align-items:center;justify-content:center;width:28px;height:28px;'
            f'margin:2px;border-radius:6px;background:{aa_colors.get(residue, "#FFFFFF")};'
            'border:1px solid rgba(15,23,42,0.08);color:#1F2937;font-weight:600;font-family:Consolas, monospace;">'
            f'{residue}</span>'
            for residue in seq
        )
        row_html.append(
            f'<div style="display:flex;align-items:center;gap:12px;margin:8px 0;">'
            f'<div style="width:96px;color:#AEB7C8;font-weight:600;">{name}</div>'
            f'<div>{cells}</div>'
            '</div>'
        )

    st.markdown(
        """
        <div style="margin-top:16px;padding:16px 18px;border:1px solid #232938;border-radius:12px;background:linear-gradient(180deg,#151922 0%,#12161E 100%);">
            <div style="color:#D6DCEA;font-weight:600;margin-bottom:6px;">Preview</div>
            <div style="color:#8B95A7;font-size:0.92rem;margin-bottom:10px;">Fake alignment preview using the current amino acid colors.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("".join(row_html), unsafe_allow_html=True)

# ── Parsers ───────────────────────────────────────────────────────────────────
def parse_fasta(content: str):
    sequences = []
    current_name, current_seq = None, []
    for line in content.splitlines():
        line = line.strip()
        if line.startswith('>'):
            if current_name is not None:
                sequences.append((current_name, ''.join(current_seq)))
            current_name = line[1:]
            current_seq = []
        elif line:
            current_seq.append(line)
    if current_name is not None:
        sequences.append((current_name, ''.join(current_seq)))
    return sequences

def parse_conservation(content: str):
    data = {}
    reader = csv.reader(content.splitlines())
    for row in reader:
        if row:
            data[row[0]] = row[1:]
    return data

def get_positions(human_seq: str):
    positions, pos = [], 1
    for ch in human_seq:
        if ch == '-':
            positions.append(None)
        else:
            positions.append(pos)
            pos += 1
    return positions

# ── Core Excel builder ────────────────────────────────────────────────────────
def build_excel(sequences, conservation, aa_colors, domains, reference_name=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alignment"

    # Use selected reference sequence for position numbering.
    # Fallback: human/homo match, then the first sequence.
    reference_seq = None
    if reference_name:
        reference_seq = next((seq for name, seq in sequences if name == reference_name), None)
    if reference_seq is None:
        reference_seq = next(
            (seq for name, seq in sequences if 'human' in name.lower() or 'homo' in name.lower()),
            sequences[0][1]
        )

    positions = get_positions(reference_seq)
    pos_to_col = {pos: idx for idx, pos in enumerate(positions, start=2) if pos is not None}
    max_col = len(reference_seq) + 1

    # ── Row 1: position numbers ───────────────────────────────────────────────
    for col_idx, pos in enumerate(positions, start=2):
        if pos is not None:
            ws.cell(row=1, column=col_idx, value=pos)

    # ── Sequence rows ─────────────────────────────────────────────────────────
    current_row = 2
    seq_start_row = 2
    for seq_name, seq in sequences:
        ws.cell(row=current_row, column=1, value=seq_name)
        for col_idx, char in enumerate(seq, start=2):
            ws.cell(row=current_row, column=col_idx, value=char)
        current_row += 1
    seq_end_row = current_row - 1

    # ── Conservation rows ─────────────────────────────────────────────────────
    conservation_rows = []
    if conservation:
        for metric, values in conservation.items():
            if metric.lower() == 'metric':
                continue
            ws.cell(row=current_row, column=1, value=metric)
            for col_idx, value in enumerate(values, start=2):
                try:
                    ws.cell(row=current_row, column=col_idx, value=float(value) if value else None)
                except (ValueError, TypeError):
                    ws.cell(row=current_row, column=col_idx, value=value)
            conservation_rows.append(current_row)
            current_row += 1

    # ── Domain label row ──────────────────────────────────────────────────────
    domain_row = current_row
    ws.cell(row=domain_row, column=1, value="Domains")
    ws.cell(row=domain_row, column=1).font = Font(bold=True)

    for d in domains:
        start_pos, end_pos, label = d["start"], d["end"], d["label"]
        cols = [pos_to_col[p] for p in range(start_pos, end_pos + 1) if p in pos_to_col]
        if not cols:
            continue
        mid_col = cols[len(cols) // 2]
        ws.cell(row=domain_row, column=mid_col, value=label)
        ws.cell(row=domain_row, column=mid_col).font = Font(bold=True)
        # Bold position numbers for this range
        for col in cols:
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)

    current_row += 1

    # ── Conditional formatting: amino acids ───────────────────────────────────
    cf_range = f"B{seq_start_row}:{get_column_letter(max_col)}{seq_end_row}"
    for aa, hex_color in aa_colors.items():
        argb = hex_to_argb(hex_color)
        fill = PatternFill(start_color=argb, end_color=argb, fill_type='solid')
        dxf = DifferentialStyle(fill=fill)
        rule = Rule(type='containsText', operator='containsText', dxf=dxf)
        rule.formula = [f'NOT(ISERROR(SEARCH("{aa}",B{seq_start_row})))']
        rule.text = aa
        ws.conditional_formatting.add(cf_range, rule)

    # ── Color scale: conservation rows ───────────────────────────────────────
    for row_idx in conservation_rows:
        r_range = f"B{row_idx}:{get_column_letter(max_col)}{row_idx}"
        ws.conditional_formatting.add(r_range, ColorScaleRule(
            start_type='min',        start_color='FFFFFFFF',
            mid_type='percentile',   mid_value=50, mid_color='FFFFEB84',
            end_type='max',          end_color='FF63BE7B',
        ))

    # ── Cell formatting ───────────────────────────────────────────────────────
    for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(
                horizontal='left' if cell.column == 1 else 'center',
                vertical='center'
            )

    ws.column_dimensions['A'].width = 50
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ═════════════════════════════════════════════════════════════════════════════
# UI
# ═════════════════════════════════════════════════════════════════════════════

st.title("🧬 Alignment Formatter")
st.markdown("Upload a FASTA alignment and (optionally) conservation scores to generate a color-coded Excel workbook.")

tab_files, tab_domains, tab_colors = st.tabs(["📁 Files", "📍 Domains & Highlights", "🎨 Amino Acid Colors"])

# ── Tab 1: Files ──────────────────────────────────────────────────────────────
with tab_files:
    col1, col2 = st.columns(2)
    with col1:
        aln_file = st.file_uploader("Alignment file (.fas / .aln / .fasta)", type=["fas", "aln", "fasta", "txt"])
    with col2:
        csv_file = st.file_uploader("Conservation scores (.csv)  — optional", type=["csv"])

    if aln_file:
        sequences = parse_fasta(aln_file.read().decode('utf-8'))
        st.success(f"✅ Loaded **{len(sequences)}** sequences  |  Length: **{len(sequences[0][1])}** aligned columns")
        default_reference = next((n for n, _ in sequences if 'human' in n.lower() or 'homo' in n.lower()), sequences[0][0])
        seq_names = [name for name, _ in sequences]

        # Keep selection stable across reruns if the chosen sequence still exists.
        if st.session_state.get("reference_sequence") not in seq_names:
            st.session_state.reference_sequence = default_reference

        reference_name = st.selectbox(
            "Reference sequence for numbering",
            options=seq_names,
            key="reference_sequence",
            help="Residue numbering in row 1 follows this sequence (gaps are skipped).",
        )
        st.info(f"Reference sequence (for numbering): **{reference_name}**")
        with st.expander("Preview sequences"):
            for name, seq in sequences:
                st.markdown(f"`{name}`")
    
    if csv_file:
        conservation = parse_conservation(csv_file.read().decode('utf-8'))
        metrics = [k for k in conservation if k.lower() != 'metric']
        st.success(f"✅ Loaded conservation scores: {', '.join(metrics)}")

# ── Tab 2: Domains ────────────────────────────────────────────────────────────
with tab_domains:
    st.caption("Residue positions and domain labels in this section align to the selected REFERENCE sequence from the Files tab.")
    st.markdown("""
    <div class="info-box">
    Define protein regions to <b>bold</b> and label in the output. Each domain highlights a range of residue positions based on the selected <b>REFERENCE</b> sequence.
    Single residues: set start = end.
    </div>
    """, unsafe_allow_html=True)

    # Header
    hc = st.columns([1.2, 1.2, 2, 0.6])
    hc[0].markdown("**Start residue**")
    hc[1].markdown("**End residue**")
    hc[2].markdown("**Label**")
    hc[3].markdown("**Del**")

    to_delete = []
    for i, domain in enumerate(st.session_state.domains):
        cols = st.columns([1.2, 1.2, 2, 0.6])
        st.session_state.domains[i]["start"] = cols[0].number_input(
            f"start_{i}", value=domain["start"], min_value=1, label_visibility="collapsed", key=f"ds_{i}"
        )
        st.session_state.domains[i]["end"] = cols[1].number_input(
            f"end_{i}", value=domain["end"], min_value=1, label_visibility="collapsed", key=f"de_{i}"
        )
        st.session_state.domains[i]["label"] = cols[2].text_input(
            f"label_{i}", value=domain["label"], label_visibility="collapsed", key=f"dl_{i}"
        )
        if cols[3].button("✕", key=f"del_{i}"):
            to_delete.append(i)

    for i in reversed(to_delete):
        st.session_state.domains.pop(i)

    col_add, col_clear = st.columns([1, 4])
    with col_add:
        if st.button("＋ Add domain"):
            last = st.session_state.domains[-1] if st.session_state.domains else {"end": 0}
            st.session_state.domains.append({"start": last["end"] + 1, "end": last["end"] + 10, "label": "New"})
            st.rerun()
    with col_clear:
        if st.button("Clear all domains"):
            st.session_state.domains = []
            st.rerun()

    st.markdown("---")
    st.markdown("**Paste domain list** (one per line: `start-end Label` or `pos Label`, using numbering from the selected `REFERENCE` sequence)")
    bulk = st.text_area("e.g.  `266-288 ZF1`  or  `244 K244`", height=120, label_visibility="collapsed")
    if st.button("Import from text"):
        new_domains = []
        for line in bulk.strip().splitlines():
            line = line.strip()
            if not line:
                continue
            m = re.match(r'^(\d+)-(\d+)\s+(.+)$', line)
            if m:
                new_domains.append({"start": int(m.group(1)), "end": int(m.group(2)), "label": m.group(3).strip()})
            else:
                m2 = re.match(r'^(\d+)\s+(.+)$', line)
                if m2:
                    new_domains.append({"start": int(m2.group(1)), "end": int(m2.group(1)), "label": m2.group(2).strip()})
        if new_domains:
            st.session_state.domains.extend(new_domains)
            st.success(f"Added {len(new_domains)} domains.")
            st.rerun()

# ── Tab 3: Colors ─────────────────────────────────────────────────────────────
with tab_colors:
    st.markdown("Customize the fill color for each amino acid. Changes are reflected in the generated file.")

    st.markdown("### Templates")
    st.caption("Apply a starter palette, then fine-tune individual residues below.")
    temp_col1, temp_col2, _ = st.columns([2.2, 1.2, 2])
    selected_template = temp_col1.selectbox(
        "Template color scheme",
        options=list(COLOR_TEMPLATES.keys()),
        key="selected_color_template",
    )
    if temp_col2.button("Apply template"):
        apply_color_scheme(COLOR_TEMPLATES[selected_template])
        st.rerun()

    col_reset, _ = st.columns([1, 4])
    with col_reset:
        if st.button("Reset to defaults"):
            apply_color_scheme(DEFAULT_AA_COLORS)
            st.rerun()

    for group_name, aas in AA_GROUPS.items():
        st.markdown(f"**{group_name}**")
        cols = st.columns(len(aas))
        for col, aa in zip(cols, aas):
            picked = col.color_picker(aa, key=f"color_{aa}")
            st.session_state.aa_colors[aa] = picked

    render_alignment_preview(st.session_state.aa_colors)

# ── Generate button ───────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("### Generate Excel")

if st.button("⬇️  Generate & Download", use_container_width=True):
    if not aln_file:
        st.error("Please upload an alignment file first.")
    else:
        aln_file.seek(0)
        sequences = parse_fasta(aln_file.read().decode('utf-8'))
        conservation = {}
        if csv_file:
            csv_file.seek(0)
            conservation = parse_conservation(csv_file.read().decode('utf-8'))

        with st.spinner("Building Excel file…"):
            buf = build_excel(
                sequences,
                conservation,
                st.session_state.aa_colors,
                st.session_state.domains,
                st.session_state.get("reference_sequence"),
            )

        base = aln_file.name.rsplit('.', 1)[0]
        st.download_button(
            label="📥 Download formatted alignment",
            data=buf,
            file_name=f"{base}_formatted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.success("Done! Click the button above to download your file.")
